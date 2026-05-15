# factures/exports.py
import csv
import codecs
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

class FactureExport:
    """Classe pour exporter les factures en CSV et Excel"""
    
    @staticmethod
    def exporter_csv(factures, request):
        """Exporte les factures au format CSV avec UTF-8"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="factures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        response.write(codecs.BOM_UTF8)
        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_ALL)
        
        # ✅ En-têtes avec commissions
        writer.writerow([
            'N° Facture', 'Client', 'Date facture', 'Date échéance', 
            'Type', 'Statut', 'Total HT', 'TVA', 'Total TTC',
            'Taux Commission (%)', 'Montant Commission', 'Commissionnaire',
            'Net à payer', 'Total payé', 'Reste à payer', 'Date création'
        ])
        
        for facture in factures:
            writer.writerow([
                facture.numero,
                facture.client.nom,
                facture.date_facture.strftime('%d/%m/%Y'),
                facture.date_echeance.strftime('%d/%m/%Y'),
                facture.get_type_facture_display(),
                facture.get_statut_display(),
                f"{facture.total_ht:.2f}".replace('.', ','),
                f"{facture.montant_tva:.2f}".replace('.', ','),
                f"{facture.total_ttc:.2f}".replace('.', ','),
                f"{float(facture.taux_commission):.2f}".replace('.', ','),  # ✅ Taux commission
                f"{float(facture.montant_commission):.2f}".replace('.', ','),  # ✅ Montant commission
                facture.commissionnaire or '',  # ✅ Commissionnaire
                f"{float(facture.net_a_payer):.2f}".replace('.', ','),  # ✅ Net à payer
                f"{float(facture.total_paye):.2f}".replace('.', ','),
                f"{float(facture.reste_a_payer):.2f}".replace('.', ','),
                facture.date_creation.strftime('%d/%m/%Y %H:%M'),
            ])
        
        return response
    
    @staticmethod
    def exporter_excel(factures, request):
        """Exporte les factures au format Excel avec commissions"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Factures"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
        
        # ✅ En-têtes avec commissions
        headers = [
            'N° Facture', 'Client', 'Date facture', 'Date échéance',
            'Type', 'Statut', 'Total HT', 'TVA', 'Total TTC',
            'Taux Commission (%)', 'Montant Commission', 'Commissionnaire',
            'Net à payer', 'Total payé', 'Reste à payer', 'Date création'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for row, facture in enumerate(factures, 2):
            ws.cell(row=row, column=1, value=facture.numero)
            ws.cell(row=row, column=2, value=facture.client.nom)
            ws.cell(row=row, column=3, value=facture.date_facture.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=4, value=facture.date_echeance.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=5, value=facture.get_type_facture_display())
            ws.cell(row=row, column=6, value=facture.get_statut_display())
            ws.cell(row=row, column=7, value=float(facture.total_ht))
            ws.cell(row=row, column=8, value=float(facture.montant_tva))
            ws.cell(row=row, column=9, value=float(facture.total_ttc))
            ws.cell(row=row, column=10, value=float(facture.taux_commission))  # ✅
            ws.cell(row=row, column=11, value=float(facture.montant_commission))  # ✅
            ws.cell(row=row, column=12, value=facture.commissionnaire or '')  # ✅
            ws.cell(row=row, column=13, value=float(facture.net_a_payer))  # ✅
            ws.cell(row=row, column=14, value=float(facture.total_paye))
            ws.cell(row=row, column=15, value=float(facture.reste_a_payer))
            ws.cell(row=row, column=16, value=facture.date_creation.strftime('%d/%m/%Y %H:%M'))
            
            # Format des nombres
            for col in [7, 8, 9, 10, 11, 13, 14, 15]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
        
        # Ajuster les largeurs des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="factures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response


class StatistiqueExport:
    """Classe pour exporter les statistiques avec commissions"""
    
    @staticmethod
    def exporter_statistiques_excel(factures, request):
        """Exporte les statistiques des factures au format Excel avec commissions"""
        wb = Workbook()
        
        # ========== Feuille 1: Récapitulatif ==========
        ws1 = wb.active
        ws1.title = "Récapitulatif"
        
        total_factures = factures.count()
        total_ht = sum(f.total_ht for f in factures)
        total_tva = sum(f.montant_tva for f in factures)
        total_ttc = sum(f.total_ttc for f in factures)
        total_commissions = sum(f.montant_commission for f in factures)
        total_net = sum(f.net_a_payer for f in factures)
        total_paye = sum(f.total_paye for f in factures)
        total_reste = sum(f.reste_a_payer for f in factures)
        
        # Statistiques par statut
        factures_payees = factures.filter(statut='payee').count()
        factures_impayees = factures.filter(statut='impayee').count()
        factures_envoyees = factures.filter(statut='envoyee').count()
        factures_partiel = factures.filter(statut='partiel').count()
        factures_definitive = factures.filter(statut='definitive').count()
        
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        
        ws1.cell(row=1, column=1, value="STATISTIQUES DES FACTURES AVEC COMMISSIONS")
        ws1.cell(row=1, column=1).font = title_font
        
        row = 3
        stats_data = [
            ('📊 GÉNÉRAL', ''),
            ('Nombre total de factures', f"{total_factures}"),
            ('Total HT', f"{total_ht:,.2f} FCFA"),
            ('Total TVA', f"{total_tva:,.2f} FCFA"),
            ('Total TTC', f"{total_ttc:,.2f} FCFA"),
            ('', ''),
            ('💰 COMMISSIONS', ''),
            ('Total des commissions', f"{total_commissions:,.2f} FCFA"),
            ('Net à payer (TTC - Commissions)', f"{total_net:,.2f} FCFA"),
            ('Commission moyenne par facture', f"{(total_commissions/total_factures if total_factures > 0 else 0):,.2f} FCFA"),
            ('', ''),
            ('💳 PAIEMENTS', ''),
            ('Total payé', f"{total_paye:,.2f} FCFA"),
            ('Reste à payer', f"{total_reste:,.2f} FCFA"),
            ('Taux de recouvrement', f"{(total_paye/total_ttc*100 if total_ttc > 0 else 0):.1f}%"),
            ('', ''),
            ('📌 PAR STATUT', ''),
            ('Définitives', f"{factures_definitive}"),
            ('Payées', f"{factures_payees} ({factures_payees/total_factures*100 if total_factures > 0 else 0:.1f}%)"),
            ('Impayées', f"{factures_impayees} ({factures_impayees/total_factures*100 if total_factures > 0 else 0:.1f}%)"),
            ('Envoyées', f"{factures_envoyees} ({factures_envoyees/total_factures*100 if total_factures > 0 else 0:.1f}%)"),
            ('Partiellement payées', f"{factures_partiel} ({factures_partiel/total_factures*100 if total_factures > 0 else 0:.1f}%)"),
        ]
        
        for label, value in stats_data:
            if label and not value:  # Titre de section
                ws1.cell(row=row, column=1, value=label).font = header_font
            else:
                ws1.cell(row=row, column=1, value=label)
                if value:
                    ws1.cell(row=row, column=2, value=value)
            row += 1
        
        ws1.column_dimensions['A'].width = 35
        ws1.column_dimensions['B'].width = 35
        
        # ========== Feuille 2: Par Commissionnaire ==========
        ws2 = wb.create_sheet("Par Commissionnaire")
        
        from collections import defaultdict
        par_comm = defaultdict(lambda: {'nb': 0, 'ttc': 0, 'commission': 0, 'net': 0})
        
        for f in factures:
            comm = f.commissionnaire or 'Sans commissionnaire'
            par_comm[comm]['nb'] += 1
            par_comm[comm]['ttc'] += f.total_ttc
            par_comm[comm]['commission'] += f.montant_commission
            par_comm[comm]['net'] += f.net_a_payer
        
        headers2 = ['Commissionnaire', 'Nb factures', 'Total TTC', 'Total Commission', 'Net à payer', 'Taux moyen']
        for col, header in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=header).font = header_font
            ws2.cell(row=1, column=col).fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
            ws2.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        
        row = 2
        for comm, data in sorted(par_comm.items(), key=lambda x: x[1]['commission'], reverse=True):
            taux_moyen = (data['commission'] / data['ttc'] * 100) if data['ttc'] > 0 else 0
            ws2.cell(row=row, column=1, value=comm)
            ws2.cell(row=row, column=2, value=data['nb'])
            ws2.cell(row=row, column=3, value=float(data['ttc']))
            ws2.cell(row=row, column=4, value=float(data['commission']))
            ws2.cell(row=row, column=5, value=float(data['net']))
            ws2.cell(row=row, column=6, value=f"{taux_moyen:.2f}%")
            
            for col in [3, 4, 5]:
                ws2.cell(row=row, column=col).number_format = '#,##0.00'
            row += 1
        
        for col in range(1, 7):
            ws2.column_dimensions[get_column_letter(col)].width = 20
        
        # ========== Feuille 3: Liste détaillée ==========
        ws3 = wb.create_sheet("Liste détaillée")
        
        headers3 = ['N° Facture', 'Client', 'Date', 'Statut', 'Total TTC', 'Commission', 'Net à payer', 'Commissionnaire']
        for col, header in enumerate(headers3, 1):
            ws3.cell(row=1, column=col, value=header).font = header_font
            ws3.cell(row=1, column=col).fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
            ws3.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        
        for row, facture in enumerate(factures, 2):
            ws3.cell(row=row, column=1, value=facture.numero)
            ws3.cell(row=row, column=2, value=facture.client.nom)
            ws3.cell(row=row, column=3, value=facture.date_facture.strftime('%d/%m/%Y'))
            ws3.cell(row=row, column=4, value=facture.get_statut_display())
            ws3.cell(row=row, column=5, value=float(facture.total_ttc))
            ws3.cell(row=row, column=6, value=float(facture.montant_commission))
            ws3.cell(row=row, column=7, value=float(facture.net_a_payer))
            ws3.cell(row=row, column=8, value=facture.commissionnaire or '')
            
            for col in [5, 6, 7]:
                ws3.cell(row=row, column=col).number_format = '#,##0.00'
        
        for col in range(1, 9):
            ws3.column_dimensions[get_column_letter(col)].width = 18
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="statistiques_factures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response


# ✅ Fonction export comptabilité avec commissions
def export_comptabilite(request):
    """Export pour la comptabilité avec les commissions"""
    from factures.models import Facture
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Comptabilité"
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    
    # ✅ En-têtes avec commissions
    headers = [
        'Date facture', 'N° Facture', 'Client', 'Total HT', 
        'TVA', 'Total TTC', 'Commission (%)', 'Montant Commission', 
        'Net à payer', 'Statut', 'Date paiement'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    factures = Facture.objects.filter(type_facture='definitive')
    
    for row, facture in enumerate(factures, 2):
        ws.cell(row=row, column=1, value=facture.date_facture.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=facture.numero)
        ws.cell(row=row, column=3, value=facture.client.nom)
        ws.cell(row=row, column=4, value=float(facture.total_ht))
        ws.cell(row=row, column=5, value=float(facture.montant_tva))
        ws.cell(row=row, column=6, value=float(facture.total_ttc))
        ws.cell(row=row, column=7, value=float(facture.taux_commission))
        ws.cell(row=row, column=8, value=float(facture.montant_commission))
        ws.cell(row=row, column=9, value=float(facture.net_a_payer))
        ws.cell(row=row, column=10, value=facture.get_statut_display())
        
        dernier_paiement = facture.paiements.filter(statut='confirme').order_by('-date_paiement').first()
        ws.cell(row=row, column=11, value=dernier_paiement.date_paiement.strftime('%d/%m/%Y') if dernier_paiement else '')
        
        for col in [4, 5, 6, 7, 8, 9]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="export_comptabilite_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response