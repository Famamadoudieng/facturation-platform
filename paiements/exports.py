# paiements/exports.py
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

class PaiementExport:
    """Classe pour exporter les paiements en CSV et Excel"""
    
    @staticmethod
    def exporter_csv(paiements, request):
        """Exporte les paiements au format CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="paiements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        writer.writerow([
            'N° Paiement', 'Facture', 'Client', 'Date paiement', 
            'Mode', 'Référence', 'Montant', 'Type', 'Statut', 
            'Notes', 'Date création'
        ])
        
        # Données
        for paiement in paiements:
            writer.writerow([
                paiement.numero_paiement,
                paiement.facture.numero,
                paiement.facture.client.nom,
                paiement.date_paiement.strftime('%d/%m/%Y'),
                paiement.get_mode_paiement_display(),
                paiement.reference or '',
                f"{paiement.montant:.2f}",
                "Acompte" if paiement.facture.type_facture == 'proforma' else "Paiement",
                paiement.get_statut_display(),
                paiement.notes or '',
                paiement.created_at.strftime('%d/%m/%Y %H:%M'),
            ])
        
        return response
    
    @staticmethod
    def exporter_excel(paiements, request):
        """Exporte les paiements au format Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Paiements"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = [
            'N° Paiement', 'Facture', 'Client', 'Date paiement',
            'Mode', 'Référence', 'Montant', 'Type', 'Statut', 'Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Données
        for row, paiement in enumerate(paiements, 2):
            ws.cell(row=row, column=1, value=paiement.numero_paiement)
            ws.cell(row=row, column=2, value=paiement.facture.numero)
            ws.cell(row=row, column=3, value=paiement.facture.client.nom)
            ws.cell(row=row, column=4, value=paiement.date_paiement.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=5, value=paiement.get_mode_paiement_display())
            ws.cell(row=row, column=6, value=paiement.reference or '')
            ws.cell(row=row, column=7, value=float(paiement.montant))
            ws.cell(row=row, column=8, value="Acompte" if paiement.facture.type_facture == 'proforma' else "Paiement")
            ws.cell(row=row, column=9, value=paiement.get_statut_display())
            ws.cell(row=row, column=10, value=paiement.notes or '')
        
        # Ajuster les largeurs
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="paiements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response


# paiements/exports.py - Version simplifiée sans fusion

class StatistiquePaiementExport:
    @staticmethod
    def exporter_statistiques_excel(paiements, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistiques"
        
        from django.db.models import Sum
        
        total_paiements = paiements.count()
        total_montant = paiements.filter(statut='confirme').aggregate(total=Sum('montant'))['total'] or 0
        
        # Écrire ligne par ligne sans fusion
        ws.cell(row=1, column=1, value="STATISTIQUES DES PAIEMENTS")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=3, column=1, value="Nombre total de paiements")
        ws.cell(row=3, column=2, value=total_paiements)
        
        ws.cell(row=4, column=1, value="Montant total encaissé")
        ws.cell(row=4, column=2, value=f"{total_montant:,.2f} FCFA")
        
        # Feuille 2: Liste détaillée
        ws2 = wb.create_sheet("Liste détaillée")
        
        headers = ['N° Paiement', 'Facture', 'Client', 'Date', 'Mode', 'Montant', 'Statut']
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header)
            ws2.cell(row=1, column=col).font = Font(bold=True)
        
        for row, paiement in enumerate(paiements, 2):
            ws2.cell(row=row, column=1, value=paiement.numero_paiement)
            ws2.cell(row=row, column=2, value=paiement.facture.numero)
            ws2.cell(row=row, column=3, value=paiement.facture.client.nom)
            ws2.cell(row=row, column=4, value=paiement.date_paiement.strftime('%d/%m/%Y'))
            ws2.cell(row=row, column=5, value=paiement.get_mode_paiement_display())
            ws2.cell(row=row, column=6, value=float(paiement.montant))
            ws2.cell(row=row, column=7, value=paiement.get_statut_display())
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="statistiques_paiements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response